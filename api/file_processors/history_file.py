import openpyxl
from datetime import datetime


class HistoryFileWriter:

    def __init__(self, data):
        self.data = data

    def write(self, response):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Report'

        key_item = 'Localization key'
        tags_item = 'Tags'
        comment_item = 'Comments'
        language_item = 'Language'
        translation_item = 'Translation'
        update_item = 'Updated at'

        header = [key_item, language_item,
                  translation_item, tags_item, comment_item, update_item]

        indexes = {k: v + 1 for v, k in enumerate(header)}

        for idx in range(len(header)):
            cell = ws.cell(row=1, column=idx+1)
            cell.value = header[idx]

        row = 2

        for item in self.data:
            ws.cell(
                row=row,
                column=indexes[key_item],
                value=item.token.token
            )
            ws.cell(
                row=row,
                column=indexes[language_item],
                value=item.language
            )
            ws.cell(
                row=row,
                column=indexes[translation_item],
                value=item.new_value
            )
            ws.cell(
                row=row,
                column=indexes[tags_item],
                value=','.join([tag.tag for tag in item.token.tags.all()])
            )
            ws.cell(
                row=row,
                column=indexes[comment_item],
                value=item.token.comment
            )
            ws.cell(
                row=row,
                column=indexes[update_item],
                value=item.updated_at.strftime('%Y-%m-%d %H:%M')
            )
            row += 1

        wb.save(response)
