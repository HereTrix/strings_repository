from django.http import HttpResponse
import openpyxl
import tempfile


class FileConstants:
    key_item = 'Localization key'
    tags_item = 'Tags'
    comment_item = 'Comments'
    translation_item = 'Translation'


class ExcelFileWriter:

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.has_data = False

    def append(self, records, code):
        if self.has_data:
            ws = self.wb.create_sheet(code)
        else:
            ws = self.wb.active
            ws.title = code
            self.has_data = True

        header = [FileConstants.key_item,
                  FileConstants.translation_item,
                  FileConstants.tags_item,
                  FileConstants.comment_item]
        indexes = {k: v + 1 for v, k in enumerate(header)}

        for idx in range(len(header)):
            cell = ws.cell(row=1, column=idx+1)
            cell.value = header[idx]

        row = 2

        for record in records:
            ws.cell(
                row=row,
                column=indexes[FileConstants.key_item],
                value=record.token
            )
            ws.cell(
                row=row,
                column=indexes[FileConstants.translation_item],
                value=record.translation
            )
            ws.cell(
                row=row,
                column=indexes[FileConstants.comment_item],
                value=record.comment
            )
            ws.cell(
                row=row,
                column=indexes[FileConstants.tags_item],
                value=','.join(record.tags)
            )
            row += 1

    def http_response(self):
        with tempfile.NamedTemporaryFile() as tmp:
            self.wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type='application/ms-excel'
        )
        response['Content-Disposition'] = 'attachment; filename=translations.xlsx'
        return response


class ExcelSingleSheetFileWriter:

    def __init__(self) -> None:
        self.records = {}
        self.languages = []

    def append(self, records, code):
        for item in records:
            record = self.records.get(item.token)
            if not record:
                record = {
                    FileConstants.key_item: item.token,
                    FileConstants.comment_item: item.comment,
                    FileConstants.tags_item: ','.join(item.tags)
                }
            record[code] = item.translation
            self.records[item.token] = record
        self.languages.append(code)

    def http_response(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        header = [FileConstants.key_item] + \
            self.languages + \
            [FileConstants.tags_item, FileConstants.comment_item]
        indexes = {k: v + 1 for v, k in enumerate(header)}

        for idx in range(len(header)):
            cell = ws.cell(row=1, column=idx+1)
            cell.value = header[idx]

        row = 2

        for key in self.records.keys():
            record = self.records[key]
            ws.cell(
                row=row,
                column=indexes[FileConstants.key_item],
                value=key
            )
            ws.cell(
                row=row,
                column=indexes[FileConstants.comment_item],
                value=record[FileConstants.comment_item]
            )
            ws.cell(
                row=row,
                column=indexes[FileConstants.tags_item],
                value=record[FileConstants.tags_item]
            )

            for code in self.languages:
                ws.cell(
                    row=row,
                    column=indexes[code],
                    value=record[code]
                )

            row += 1

        with tempfile.NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(
            content=stream,
            content_type='application/ms-excel'
        )
        response['Content-Disposition'] = 'attachment; filename=translations.xlsx'
        return response
